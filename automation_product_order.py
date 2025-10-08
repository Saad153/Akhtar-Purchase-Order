import pandas as pd
import re
import openpyxl
import io
from datetime import datetime

def read_file(file_content, filename):
    """Read the file content from bytes and return DataFrame"""
    try:
        file_buffer = io.BytesIO(file_content)
        header = file_content[:10].decode('utf-8', errors='ignore')

        if header.startswith('<!DOCTYPE') or header.startswith('<html'):
            dfs = pd.read_html(file_buffer)
            df = dfs[0]
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)
            return df
        else:
            try:
                if filename.endswith('.xls'):
                    return pd.read_excel(file_buffer, engine='xlrd')
                else:
                    return pd.read_excel(file_buffer, engine='openpyxl')
            except Exception as e:
                raise Exception(f"Failed to read file: {e}. Please ensure the file is a valid Excel file.")
    except Exception as e:
        raise Exception(f"Error processing file: {e}")

def format_date(date_str):
    """Convert date from dd/mm/yyyy to dd-MMM-yyyy format"""
    try:
        # Parse the input date
        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
        # Convert to desired format
        return date_obj.strftime("%d-%b-%Y")
    except ValueError as e:
        raise ValueError(f"Invalid date format. Expected dd/mm/yyyy, got: {date_str}")


def process_excel_files(po_file_path, order_form_path):
    """
    Process the PO file and order form

    Args:
        po_file_path (str): Path to PO Excel file
        order_form_path (str): Path to order form Excel file

    Returns:
        tuple: (processed_order_form_bytes, list_of_labels_and_quantities)
    """
    try:
        # Read PO file
        with open(po_file_path, 'rb') as f:
            content = f.read()
        df = read_file(content, po_file_path)

        # Regex to match label patterns
        pattern = re.compile(r'LABEL\s+([A-Z]-?\d+|[A-Z]\d+|[A-Z]\s?\d+)', re.IGNORECASE)

        # Detect Name and Quantity columns
        name_column = next((col for col in df.columns if 'name' in str(col).lower()), None)
        qty_column = next((col for col in df.columns if 'qty' in str(col).lower()), None)

        if not (name_column and qty_column):
            raise ValueError("Required columns not found in PO file")

        # Extract labels and quantities
        labels_and_quantities = []
        for _, row in df.iterrows():
            matches = pattern.findall(str(row[name_column]))
            if matches:
                for label in matches:
                    labels_and_quantities.append({
                        "label": label,
                        "quantity": row[qty_column]
                    })

        po_number = df['Po No'].iloc[0] if 'Po No' in df.columns else None
        
        if not po_number:
            raise ValueError("Could not find PO number in file")

        # Process Order Form
        with open(order_form_path, 'rb') as f:
            order_content = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(order_content))
        ws = wb.active

        # Locate "Main Labels" row
        main_labels_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if isinstance(cell_value, str) and "Main Labels" in cell_value:
                main_labels_row = row
            elif isinstance(cell_value, str) and "PO NUMBER:" in cell_value:
                po_num_row = row
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=row, column=col).value == "PO NUMBER:":
                        po_num_col = col
                        break
            elif isinstance(cell_value, str) and "Order Date : " in cell_value:
                ord_date_row = row
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=row, column=col).value == "Order Date : ":
                        ord_date_col = col
                        break

            if main_labels_row and po_num_row and ord_date_row:
                break

        if main_labels_row is None:
            raise ValueError("Main Labels section not found in order form")

        # Locate Item# and Quantity columns
        item_col, qty_col = None, None
        for col in range(1, ws.max_column + 1):            
            header = ws.cell(row=main_labels_row + 1, column=col).value
            if header:
                if "Item#" in str(header):
                    item_col = col
                elif "Quantity" in str(header):
                    qty_col = col


        if not (item_col and qty_col):
            raise ValueError("Could not find Item# or Quantity columns")

        ord_date = datetime.now().strftime("%d/%m/%Y")
        formatted_date = format_date(ord_date)
        
        # Fill in labels and quantities
        current_row = main_labels_row + 1
        for item in labels_and_quantities:
            ws.cell(row=current_row, column=item_col + 1).value = item["label"]
            ws.cell(row=current_row, column=qty_col + 1).value = item["quantity"]    
            current_row += 1

        ws.cell(row=po_num_row, column=po_num_col + 1).value = po_number
        ws.cell(row=ord_date_row, column=ord_date_col + 1).value = formatted_date

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        output_filename = f"product_order_output_{timestamp}.xlsx"

        # Save file directly
        wb.save(output_filename)
        # # Save to buffer
        # output = io.BytesIO()
        # wb.save(output)
        # output.seek(0)

        return output_filename, labels_and_quantities

    except Exception as e:
        raise RuntimeError(f"Error processing Excel files: {e}")

